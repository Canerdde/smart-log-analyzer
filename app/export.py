"""
Export modülü - PDF, CSV, Excel, JSON ve XML rapor oluşturma
"""

import io
import csv
import json
import xml.etree.ElementTree as ET
from typing import List, Dict, Any
from datetime import datetime
from fastapi.responses import StreamingResponse, Response

# PDF export (opsiyonel)
try:
    from reportlab.lib.pagesizes import letter, A4
    from reportlab.lib import colors
    from reportlab.lib.units import inch
    from reportlab.platypus import (
        SimpleDocTemplate,
        Table,
        TableStyle,
        Paragraph,
        Spacer,
    )
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.enums import TA_CENTER, TA_LEFT

    REPORTLAB_AVAILABLE = True
except ImportError:
    REPORTLAB_AVAILABLE = False

# Excel export (opsiyonel)
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

from app.models import LogAnalysis, LogEntry, LogFile
from sqlalchemy.orm import Session


def export_analysis_to_csv(
    analysis: LogAnalysis, log_file: LogFile, entries: List[LogEntry]
) -> Response:
    """
    Analiz sonuçlarını CSV formatında export et

    Args:
        analysis: LogAnalysis objesi
        log_file: LogFile objesi
        entries: LogEntry listesi

    Returns:
        CSV dosyası Response
    """
    output = io.StringIO()
    writer = csv.writer(output)

    # Başlık bilgileri
    writer.writerow(["Log Analiz Raporu"])
    writer.writerow(["Dosya Adı", log_file.filename])
    writer.writerow(["Yüklenme Tarihi", log_file.uploaded_at])
    writer.writerow(["Toplam Satır", log_file.total_lines])
    writer.writerow([])

    # İstatistikler
    writer.writerow(["İSTATİSTİKLER"])
    writer.writerow(["Toplam Giriş", analysis.total_entries])
    writer.writerow(["Hata Sayısı", analysis.error_count])
    writer.writerow(["Uyarı Sayısı", analysis.warning_count])
    writer.writerow(["Bilgi Sayısı", analysis.info_count])
    writer.writerow(["Debug Sayısı", analysis.debug_count])
    writer.writerow([])

    # En sık hatalar
    if analysis.top_errors:
        writer.writerow(["EN SIK TEKRAR EDEN HATALAR"])
        writer.writerow(["Mesaj", "Sayı", "Yüzde"])
        for error in analysis.top_errors:
            writer.writerow(
                [
                    error.get("message", "")[:100],
                    error.get("count", 0),
                    f"{error.get('percentage', 0)}%",
                ]
            )
        writer.writerow([])

    # Log girişleri
    writer.writerow(["LOG GİRİŞLERİ"])
    writer.writerow(["Satır No", "Seviye", "Zaman", "Mesaj"])
    for entry in entries[:1000]:  # İlk 1000 satır
        writer.writerow(
            [
                entry.line_number,
                entry.log_level or "",
                entry.timestamp.isoformat() if entry.timestamp else "",
                entry.message[:200],  # İlk 200 karakter
            ]
        )

    output.seek(0)

    return Response(
        content=output.getvalue(),
        media_type="text/csv; charset=utf-8",
        headers={
            "Content-Disposition": f'attachment; filename="log_analysis_{log_file.id}.csv"'
        },
    )


def export_analysis_to_pdf(
    analysis: LogAnalysis, log_file: LogFile, entries: List[LogEntry]
) -> StreamingResponse:
    """
    Analiz sonuçlarını PDF formatında export et

    Args:
        analysis: LogAnalysis objesi
        log_file: LogFile objesi
        entries: LogEntry listesi

    Returns:
        PDF dosyası StreamingResponse
    """
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4)
    story = []
    styles = getSampleStyleSheet()

    # Özel stiller
    title_style = ParagraphStyle(
        "CustomTitle",
        parent=styles["Heading1"],
        fontSize=24,
        textColor=colors.HexColor("#1a1a1a"),
        spaceAfter=30,
        alignment=TA_CENTER,
    )

    heading_style = ParagraphStyle(
        "CustomHeading",
        parent=styles["Heading2"],
        fontSize=16,
        textColor=colors.HexColor("#2c3e50"),
        spaceAfter=12,
        spaceBefore=20,
    )

    # Başlık
    story.append(Paragraph("LOG ANALİZ RAPORU", title_style))
    story.append(Spacer(1, 0.3 * inch))

    # Dosya bilgileri
    story.append(Paragraph("Dosya Bilgileri", heading_style))
    file_data = [
        ["Dosya Adı", log_file.filename],
        [
            "Yüklenme Tarihi",
            (
                log_file.uploaded_at.strftime("%Y-%m-%d %H:%M:%S")
                if log_file.uploaded_at
                else ""
            ),
        ],
        ["Toplam Satır", str(log_file.total_lines)],
        ["Durum", log_file.status],
    ]
    file_table = Table(file_data, colWidths=[2 * inch, 4 * inch])
    file_table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (0, -1), colors.grey),
                ("TEXTCOLOR", (0, 0), (0, -1), colors.whitesmoke),
                ("ALIGN", (0, 0), (-1, -1), "LEFT"),
                ("FONTNAME", (0, 0), (0, -1), "Helvetica-Bold"),
                ("FONTSIZE", (0, 0), (-1, -1), 10),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 12),
                ("BACKGROUND", (1, 0), (1, -1), colors.beige),
                ("GRID", (0, 0), (-1, -1), 1, colors.black),
            ]
        )
    )
    story.append(file_table)
    story.append(Spacer(1, 0.3 * inch))

    # İstatistikler
    story.append(Paragraph("İstatistikler", heading_style))
    stats_data = [
        ["Metrik", "Değer"],
        ["Toplam Giriş", str(analysis.total_entries)],
        ["Hata Sayısı", str(analysis.error_count)],
        ["Uyarı Sayısı", str(analysis.warning_count)],
        ["Bilgi Sayısı", str(analysis.info_count)],
        ["Debug Sayısı", str(analysis.debug_count)],
    ]
    stats_table = Table(stats_data, colWidths=[3 * inch, 3 * inch])
    stats_table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#34495e")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
                ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("FONTSIZE", (0, 0), (-1, 0), 12),
                ("BOTTOMPADDING", (0, 0), (-1, 0), 12),
                ("BACKGROUND", (0, 1), (-1, -1), colors.HexColor("#ecf0f1")),
                ("GRID", (0, 0), (-1, -1), 1, colors.black),
                (
                    "ROWBACKGROUNDS",
                    (0, 1),
                    (-1, -1),
                    [colors.white, colors.HexColor("#f8f9fa")],
                ),
            ]
        )
    )
    story.append(stats_table)
    story.append(Spacer(1, 0.3 * inch))

    # En sık hatalar
    if analysis.top_errors:
        story.append(Paragraph("En Sık Tekrar Eden Hatalar", heading_style))
        error_data = [["#", "Mesaj", "Sayı", "Yüzde"]]
        for idx, error in enumerate(analysis.top_errors[:10], 1):
            error_data.append(
                [
                    str(idx),
                    error.get("message", "")[:60],
                    str(error.get("count", 0)),
                    f"{error.get('percentage', 0):.2f}%",
                ]
            )
        error_table = Table(
            error_data, colWidths=[0.5 * inch, 3.5 * inch, 1 * inch, 1 * inch]
        )
        error_table.setStyle(
            TableStyle(
                [
                    ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#e74c3c")),
                    ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
                    ("ALIGN", (0, 0), (-1, -1), "LEFT"),
                    ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                    ("FONTSIZE", (0, 0), (-1, 0), 10),
                    ("FONTSIZE", (0, 1), (-1, -1), 9),
                    ("BOTTOMPADDING", (0, 0), (-1, -1), 8),
                    ("BACKGROUND", (0, 1), (-1, -1), colors.white),
                    ("GRID", (0, 0), (-1, -1), 1, colors.black),
                ]
            )
        )
        story.append(error_table)
        story.append(Spacer(1, 0.3 * inch))

    # AI yorumu
    if analysis.ai_comment:
        story.append(Paragraph("AI Analiz Yorumu", heading_style))
        story.append(Paragraph(analysis.ai_comment, styles["Normal"]))
        story.append(Spacer(1, 0.2 * inch))

    # Log örnekleri (ilk 50)
    story.append(Paragraph("Log Giriş Örnekleri (İlk 50)", heading_style))
    sample_entries = entries[:50]
    entry_data = [["Satır", "Seviye", "Zaman", "Mesaj"]]
    for entry in sample_entries:
        timestamp_str = entry.timestamp.strftime("%H:%M:%S") if entry.timestamp else ""
        entry_data.append(
            [
                str(entry.line_number),
                entry.log_level or "",
                timestamp_str,
                (
                    entry.message[:40] + "..."
                    if len(entry.message) > 40
                    else entry.message
                ),
            ]
        )

    entry_table = Table(
        entry_data, colWidths=[0.6 * inch, 0.8 * inch, 1 * inch, 3.6 * inch]
    )
    entry_table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#3498db")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
                ("ALIGN", (0, 0), (0, -1), "CENTER"),  # Satır numarası
                ("ALIGN", (1, 0), (1, -1), "CENTER"),  # Seviye
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("FONTSIZE", (0, 0), (-1, 0), 9),
                ("FONTSIZE", (0, 1), (-1, -1), 8),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
                ("BACKGROUND", (0, 1), (-1, -1), colors.white),
                ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
                (
                    "ROWBACKGROUNDS",
                    (0, 1),
                    (-1, -1),
                    [colors.white, colors.HexColor("#f8f9fa")],
                ),
            ]
        )
    )
    story.append(entry_table)

    # Alt bilgi
    story.append(Spacer(1, 0.3 * inch))
    story.append(
        Paragraph(
            f"Rapor Oluşturulma Tarihi: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}",
            styles["Normal"],
        )
    )

    # PDF oluştur
    doc.build(story)
    buffer.seek(0)

    return StreamingResponse(
        io.BytesIO(buffer.read()),
        media_type="application/pdf",
        headers={
            "Content-Disposition": f'attachment; filename="log_analysis_{log_file.id}.pdf"'
        },
    )


def export_analysis_to_excel(
    analysis: LogAnalysis, log_file: LogFile, entries: List[LogEntry]
) -> Response:
    """
    Analiz sonuçlarını Excel formatında export et

    Args:
        analysis: LogAnalysis objesi
        log_file: LogFile objesi
        entries: LogEntry listesi

    Returns:
        Excel dosyası Response
    """
    if not OPENPYXL_AVAILABLE:
        raise ImportError("openpyxl paketi bulunamadı. Excel export için gerekli.")

    wb = Workbook()
    ws = wb.active
    ws.title = "Log Analiz Raporu"

    # Stiller
    header_fill = PatternFill(
        start_color="667eea", end_color="667eea", fill_type="solid"
    )
    header_font = Font(bold=True, color="FFFFFF", size=12)
    title_font = Font(bold=True, size=16)
    border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )
    center_align = Alignment(horizontal="center", vertical="center")

    row = 1

    # Başlık
    ws.merge_cells(f"A{row}:D{row}")
    cell = ws[f"A{row}"]
    cell.value = "LOG ANALİZ RAPORU"
    cell.font = title_font
    cell.alignment = center_align
    row += 2

    # Dosya bilgileri
    ws[f"A{row}"] = "Dosya Adı:"
    ws[f"B{row}"] = log_file.filename
    ws[f"A{row}"].font = Font(bold=True)
    row += 1

    ws[f"A{row}"] = "Yüklenme Tarihi:"
    ws[f"B{row}"] = (
        log_file.uploaded_at.strftime("%Y-%m-%d %H:%M:%S")
        if log_file.uploaded_at
        else ""
    )
    ws[f"A{row}"].font = Font(bold=True)
    row += 1

    ws[f"A{row}"] = "Toplam Satır:"
    ws[f"B{row}"] = log_file.total_lines
    ws[f"A{row}"].font = Font(bold=True)
    row += 1

    ws[f"A{row}"] = "Durum:"
    ws[f"B{row}"] = log_file.status
    ws[f"A{row}"].font = Font(bold=True)
    row += 2

    # İstatistikler
    ws[f"A{row}"] = "İSTATİSTİKLER"
    ws[f"A{row}"].font = title_font
    ws.merge_cells(f"A{row}:B{row}")
    row += 1

    stats_data = [
        ["Metrik", "Değer"],
        ["Toplam Giriş", analysis.total_entries],
        ["Hata Sayısı", analysis.error_count],
        ["Uyarı Sayısı", analysis.warning_count],
        ["Bilgi Sayısı", analysis.info_count],
        ["Debug Sayısı", analysis.debug_count],
    ]

    for stat_row in stats_data:
        ws[f"A{row}"] = stat_row[0]
        ws[f"B{row}"] = stat_row[1]
        ws[f"A{row}"].font = Font(bold=True)
        if row == row - len(stats_data) + 1:  # Header row
            ws[f"A{row}"].fill = header_fill
            ws[f"A{row}"].font = header_font
            ws[f"B{row}"].fill = header_fill
            ws[f"B{row}"].font = header_font
        row += 1

    row += 1

    # En sık hatalar
    if analysis.top_errors:
        ws[f"A{row}"] = "EN SIK TEKRAR EDEN HATALAR"
        ws[f"A{row}"].font = title_font
        ws.merge_cells(f"A{row}:D{row}")
        row += 1

        # Header
        headers = ["#", "Mesaj", "Sayı", "Yüzde"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center_align
            cell.border = border
        row += 1

        # Error rows
        for idx, error in enumerate(analysis.top_errors[:20], 1):
            ws.cell(row=row, column=1).value = idx
            ws.cell(row=row, column=2).value = error.get("message", "")[:100]
            ws.cell(row=row, column=3).value = error.get("count", 0)
            ws.cell(row=row, column=4).value = f"{error.get('percentage', 0):.2f}%"

            for col in range(1, 5):
                cell = ws.cell(row=row, column=col)
                cell.border = border
                if col == 1 or col == 3 or col == 4:
                    cell.alignment = center_align
            row += 1

        row += 1

    # En sık uyarılar
    if analysis.top_warnings:
        ws[f"A{row}"] = "EN SIK TEKRAR EDEN UYARILAR"
        ws[f"A{row}"].font = title_font
        ws.merge_cells(f"A{row}:D{row}")
        row += 1

        # Header
        headers = ["#", "Mesaj", "Sayı", "Yüzde"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col)
            cell.value = header
            cell.fill = PatternFill(
                start_color="ffc107", end_color="ffc107", fill_type="solid"
            )
            cell.font = header_font
            cell.alignment = center_align
            cell.border = border
        row += 1

        # Warning rows
        for idx, warning in enumerate(analysis.top_warnings[:20], 1):
            ws.cell(row=row, column=1).value = idx
            ws.cell(row=row, column=2).value = warning.get("message", "")[:100]
            ws.cell(row=row, column=3).value = warning.get("count", 0)
            ws.cell(row=row, column=4).value = f"{warning.get('percentage', 0):.2f}%"

            for col in range(1, 5):
                cell = ws.cell(row=row, column=col)
                cell.border = border
                if col == 1 or col == 3 or col == 4:
                    cell.alignment = center_align
            row += 1

        row += 1

    # AI yorumu
    if analysis.ai_comment:
        ws[f"A{row}"] = "AI ANALİZ YORUMU"
        ws[f"A{row}"].font = title_font
        ws.merge_cells(f"A{row}:D{row}")
        row += 1

        ws.merge_cells(f"A{row}:D{row + 2}")
        cell = ws[f"A{row}"]
        cell.value = analysis.ai_comment
        cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
        row += 3

    # Log girişleri (yeni sheet)
    if entries:
        ws2 = wb.create_sheet("Log Girişleri")

        # Header
        headers = ["Satır No", "Seviye", "Zaman", "Mesaj"]
        for col, header in enumerate(headers, 1):
            cell = ws2.cell(row=1, column=col)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center_align
            cell.border = border

        # Log entries (ilk 5000 satır)
        for idx, entry in enumerate(entries[:5000], 2):
            ws2.cell(row=idx, column=1).value = entry.line_number
            ws2.cell(row=idx, column=2).value = entry.log_level or ""
            ws2.cell(row=idx, column=3).value = (
                entry.timestamp.strftime("%Y-%m-%d %H:%M:%S") if entry.timestamp else ""
            )
            ws2.cell(row=idx, column=4).value = entry.message[:200]

            for col in range(1, 5):
                cell = ws2.cell(row=idx, column=col)
                cell.border = border
                if col == 1 or col == 2:
                    cell.alignment = center_align

        # Auto-adjust column widths
        ws2.column_dimensions["A"].width = 12
        ws2.column_dimensions["B"].width = 12
        ws2.column_dimensions["C"].width = 20
        ws2.column_dimensions["D"].width = 60

    # Auto-adjust column widths (ana sheet)
    ws.column_dimensions["A"].width = 25
    ws.column_dimensions["B"].width = 30
    ws.column_dimensions["C"].width = 15
    ws.column_dimensions["D"].width = 15

    # Excel dosyasını memory'de oluştur
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    return Response(
        content=buffer.read(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f'attachment; filename="log_analysis_{log_file.id}.xlsx"'
        },
    )


def export_logs_to_json(
    log_file: LogFile,
    entries: List[LogEntry],
    include_analysis: bool = False,
    analysis: LogAnalysis = None,
) -> Response:
    """
    Log girişlerini JSON formatında export et

    Args:
        log_file: LogFile objesi
        entries: LogEntry listesi (filtrelenmiş olabilir)
        include_analysis: Analiz sonuçlarını da dahil et
        analysis: LogAnalysis objesi (include_analysis=True ise gerekli)

    Returns:
        JSON dosyası Response
    """
    data = {
        "file_info": {
            "id": log_file.id,
            "filename": log_file.filename,
            "file_size": log_file.file_size,
            "total_lines": log_file.total_lines,
            "uploaded_at": (
                log_file.uploaded_at.isoformat() if log_file.uploaded_at else None
            ),
            "status": log_file.status,
        },
        "entries": [
            {
                "line_number": entry.line_number,
                "log_level": entry.log_level,
                "timestamp": entry.timestamp.isoformat() if entry.timestamp else None,
                "message": entry.message,
                "raw_line": entry.raw_line,
            }
            for entry in entries
        ],
        "export_info": {
            "exported_at": datetime.now().isoformat(),
            "total_entries": len(entries),
        },
    }

    if include_analysis and analysis:
        data["analysis"] = {
            "total_entries": analysis.total_entries,
            "error_count": analysis.error_count,
            "warning_count": analysis.warning_count,
            "info_count": analysis.info_count,
            "debug_count": analysis.debug_count,
            "top_errors": analysis.top_errors,
            "top_warnings": analysis.top_warnings,
            "time_distribution": analysis.time_distribution,
            "ai_comment": analysis.ai_comment,
            "ai_suggestions": analysis.ai_suggestions,
            "analyzed_at": (
                analysis.analyzed_at.isoformat() if analysis.analyzed_at else None
            ),
        }

    json_str = json.dumps(data, indent=2, ensure_ascii=False)

    return Response(
        content=json_str.encode("utf-8"),
        media_type="application/json",
        headers={
            "Content-Disposition": f'attachment; filename="log_export_{log_file.id}.json"'
        },
    )


def export_logs_to_xml(
    log_file: LogFile,
    entries: List[LogEntry],
    include_analysis: bool = False,
    analysis: LogAnalysis = None,
) -> Response:
    """
    Log girişlerini XML formatında export et

    Args:
        log_file: LogFile objesi
        entries: LogEntry listesi (filtrelenmiş olabilir)
        include_analysis: Analiz sonuçlarını da dahil et
        analysis: LogAnalysis objesi (include_analysis=True ise gerekli)

    Returns:
        XML dosyası Response
    """
    root = ET.Element("log_export")

    # File info
    file_info = ET.SubElement(root, "file_info")
    ET.SubElement(file_info, "id").text = str(log_file.id)
    ET.SubElement(file_info, "filename").text = log_file.filename
    ET.SubElement(file_info, "file_size").text = str(log_file.file_size)
    ET.SubElement(file_info, "total_lines").text = str(log_file.total_lines)
    if log_file.uploaded_at:
        ET.SubElement(file_info, "uploaded_at").text = log_file.uploaded_at.isoformat()
    ET.SubElement(file_info, "status").text = log_file.status

    # Entries
    entries_elem = ET.SubElement(root, "entries")
    for entry in entries:
        entry_elem = ET.SubElement(entries_elem, "entry")
        ET.SubElement(entry_elem, "line_number").text = str(entry.line_number)
        if entry.log_level:
            ET.SubElement(entry_elem, "log_level").text = entry.log_level
        if entry.timestamp:
            ET.SubElement(entry_elem, "timestamp").text = entry.timestamp.isoformat()
        ET.SubElement(entry_elem, "message").text = entry.message
        if entry.raw_line:
            ET.SubElement(entry_elem, "raw_line").text = entry.raw_line

    # Export info
    export_info = ET.SubElement(root, "export_info")
    ET.SubElement(export_info, "exported_at").text = datetime.now().isoformat()
    ET.SubElement(export_info, "total_entries").text = str(len(entries))

    # Analysis (optional)
    if include_analysis and analysis:
        analysis_elem = ET.SubElement(root, "analysis")
        ET.SubElement(analysis_elem, "total_entries").text = str(analysis.total_entries)
        ET.SubElement(analysis_elem, "error_count").text = str(analysis.error_count)
        ET.SubElement(analysis_elem, "warning_count").text = str(analysis.warning_count)
        ET.SubElement(analysis_elem, "info_count").text = str(analysis.info_count)
        ET.SubElement(analysis_elem, "debug_count").text = str(analysis.debug_count)

        if analysis.top_errors:
            top_errors_elem = ET.SubElement(analysis_elem, "top_errors")
            for error in analysis.top_errors:
                error_elem = ET.SubElement(top_errors_elem, "error")
                ET.SubElement(error_elem, "message").text = error.get("message", "")
                ET.SubElement(error_elem, "count").text = str(error.get("count", 0))
                ET.SubElement(error_elem, "percentage").text = str(
                    error.get("percentage", 0)
                )

        if analysis.top_warnings:
            top_warnings_elem = ET.SubElement(analysis_elem, "top_warnings")
            for warning in analysis.top_warnings:
                warning_elem = ET.SubElement(top_warnings_elem, "warning")
                ET.SubElement(warning_elem, "message").text = warning.get("message", "")
                ET.SubElement(warning_elem, "count").text = str(warning.get("count", 0))
                ET.SubElement(warning_elem, "percentage").text = str(
                    warning.get("percentage", 0)
                )

        if analysis.time_distribution:
            time_dist_elem = ET.SubElement(analysis_elem, "time_distribution")
            for hour, count in analysis.time_distribution.items():
                hour_elem = ET.SubElement(time_dist_elem, "hour")
                ET.SubElement(hour_elem, "hour").text = str(hour)
                ET.SubElement(hour_elem, "count").text = str(count)

        if analysis.ai_comment:
            ET.SubElement(analysis_elem, "ai_comment").text = analysis.ai_comment

    # XML'i string'e çevir
    ET.indent(root, space="  ")
    xml_str = ET.tostring(root, encoding="utf-8", xml_declaration=True).decode("utf-8")

    return Response(
        content=xml_str.encode("utf-8"),
        media_type="application/xml",
        headers={
            "Content-Disposition": f'attachment; filename="log_export_{log_file.id}.xml"'
        },
    )
